"""
Odoo 18 BoM Overview PDF → Excel Converter
Processes each page independently to avoid y-coordinate collisions.
"""

import re
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
from collections import defaultdict

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Column x-ranges ───────────────────────────────────────────────────────────
# Derived by analysing word x0 positions in the Odoo BoM Overview PDF.
COL_BOUNDS = {
    'code_name':   (0,   235),
    'quantity':    (230, 270),
    'unit':        (268, 340),
    'ready':       (338, 395),
    'free_onhand': (393, 467),
    'avail':       (465, 552),
    'lead_time':   (548, 592),
    'route':       (580, 745),
    'bom_cost':    (728, 785),
    'prod_cost':   (783, 850),
}


def col_for(x0):
    for col, (lo, hi) in COL_BOUNDS.items():
        if lo <= x0 < hi:
            return col
    return None


def in_col(x0, col):
    lo, hi = COL_BOUNDS[col]
    return lo <= x0 < hi


def zap(s):
    """Remove zero-width characters."""
    return s.replace('\u200b', '').replace('\u200c', '').strip()


def to_float(s):
    s = zap(s).replace(',', '').replace(' ', '')
    try:
        return float(s)
    except ValueError:
        return None


def euro_to_float(s):
    s = zap(s).replace('€', '').replace(',', '').strip()
    try:
        return float(s)
    except ValueError:
        return None


HEADER_RE = re.compile(
    r'^(ready|free|lead|bom|product|produce|on|hand|time|cost|unit|'
    r'availability|route|units|overview|\d{2}/\d{2}/\d{4}|days)$', re.I
)


def is_header_row(words):
    texts = [zap(w['text']).lower() for w in words if zap(w['text'])]
    return bool(texts) and all(HEADER_RE.match(t) for t in texts)


def parse_page_words(words):
    """
    Given a sorted list of words from one page, build logical rows.
    Returns a list of raw row dicts (column → word list).
    """
    # Group by y
    rows_map = defaultdict(list)
    for w in words:
        rows_map[round(w['top'])].append(w)

    logical_rows = []
    current = None

    for y, row_words in sorted(rows_map.items()):
        row_words_s = sorted(row_words, key=lambda w: w['x0'])
        texts = [zap(w['text']) for w in row_words_s]

        if is_header_row(row_words_s):
            continue

        first_text = texts[0] if texts else ''
        first_x = row_words_s[0]['x0'] if row_words_s else 999

        is_new_product  = first_text.startswith('[') and in_col(first_x, 'code_name')
        is_subcon       = first_text.lower().startswith('subcontracting')
        is_ops          = first_text.lower() == 'operations'
        is_ops_detail   = (not is_new_product and not is_subcon and not is_ops
                           and any(re.match(r'^\d+:\d+$', t) for t in texts))
        is_date         = bool(re.match(r'^\d{2}/\d{2}/\d{4}$', first_text))

        # Date/Days continuation
        if is_date and current is not None:
            current.setdefault('avail_words', []).append(first_text)
            for w, t in zip(row_words_s, texts):
                if re.match(r'^\d+$', t) and in_col(w['x0'], 'lead_time'):
                    current.setdefault('lead_time_words', []).append(t)
            continue

        # Pure continuation (route wrap, name wrap, etc.)
        if (not is_new_product and not is_subcon and not is_ops and not is_ops_detail
                and current is not None):
            for w, t in zip(row_words_s, texts):
                if t.lower() == 'days':
                    continue
                c = col_for(w['x0'])
                if c:
                    current.setdefault(c + '_words', []).append(t)
            continue

        # New logical row
        if current is not None:
            logical_rows.append(current)

        current = {'y': y}
        for w, t in zip(row_words_s, texts):
            c = col_for(w['x0'])
            if c:
                current.setdefault(c + '_words', []).append(t)

        if is_subcon:
            current['row_type'] = 'subcontracting'
        elif is_ops:
            current['row_type'] = 'operations'
        elif is_ops_detail:
            current['row_type'] = 'operation_detail'
        else:
            current['row_type'] = 'component'

    if current is not None:
        logical_rows.append(current)

    return logical_rows


def build_row(lr):
    """Convert a raw logical row dict to a clean data dict."""
    rt = lr.get('row_type', 'component')

    code_name = ' '.join(lr.get('code_name_words', []))
    m = re.match(r'(\[[^\]]+\])\s*(.*)', code_name)
    code = m.group(1) if m else ''
    name = m.group(2).strip() if m else code_name

    qty   = to_float(' '.join(lr.get('quantity_words', [])))
    unit_raw = ' '.join(lr.get('unit_words', []))
    unit  = 'Units' if unit_raw.lower().startswith('unit') else unit_raw or 'Units'

    ready = to_float(' '.join(lr.get('ready_words', [])))

    fo_str = ' '.join(lr.get('free_onhand_words', []))
    fo_parts = fo_str.split('/')
    free_to_use = to_float(fo_parts[0]) if fo_parts else None
    on_hand     = to_float(fo_parts[1]) if len(fo_parts) > 1 else None

    avail_words  = lr.get('avail_words', [])
    availability = ' '.join(avail_words).strip()

    lead_str = ' '.join(lr.get('lead_time_words', []))
    lead_m   = re.search(r'(\d+)', lead_str)
    lead     = int(lead_m.group(1)) if lead_m else None

    route    = ' '.join(lr.get('route_words', [])).strip()
    bom_cost = euro_to_float(' '.join(lr.get('bom_cost_words', [])))
    prod_cost= euro_to_float(' '.join(lr.get('prod_cost_words', [])))

    if rt == 'subcontracting' and not name:
        name = route

    return {
        'code': code,
        'name': name,
        'quantity': qty,
        'qty_unit': unit if rt == 'component' else '',
        'ready_to_produce': ready,
        'free_to_use': free_to_use,
        'on_hand': on_hand,
        'availability': availability,
        'lead_time_days': lead,
        'route': route,
        'bom_cost': bom_cost,
        'product_cost': prod_cost,
        'row_type': rt,
    }


def extract_rows(pdf_path: str):
    title = ''
    all_logical = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words(x_tolerance=3, y_tolerance=3)
            if not words:
                continue
            if page_num == 0:
                title_words = [w['text'] for w in words if w['top'] < 160]
                title = ' '.join(title_words)
            page_rows = parse_page_words(words)
            all_logical.extend(page_rows)

    rows = [build_row(lr) for lr in all_logical]

    # Filter out junk rows:
    # - Component rows that have neither a product code NOR any cost data
    # - Rows whose name contains obvious header words mixed in
    def is_valid(r):
        if r['row_type'] != 'component':
            return True
        has_code = bool(r['code'])
        has_cost = r['bom_cost'] is not None or r['product_cost'] is not None
        has_stock = r['free_to_use'] is not None or r['on_hand'] is not None
        junk_name = any(kw in (r['name'] or '').lower()
                        for kw in ['product quantity', 'availability', 'lead time'])
        return has_code and (has_cost or has_stock) and not junk_name

    rows = [r for r in rows if is_valid(r)]
    return title, rows


# ── Excel writer ──────────────────────────────────────────────────────────────

HEADER_DARK = '1F3864'
AVAIL_GREEN = 'E2EFDA'
AVAIL_RED   = 'FCE4D6'
AVAIL_EST   = 'FFF2CC'
STRIPE      = 'F5F5F5'
SUBCON_BG   = 'D9E1F2'
OPS_BG      = 'FFF2CC'

COLS = [
    ('Product Code',      16),
    ('Product Name',      40),
    ('Quantity',          10),
    ('Unit',               8),
    ('Ready to Produce',  16),
    ('Free to Use',       14),
    ('On Hand',           13),
    ('Availability',      17),
    ('Lead Time (Days)',  16),
    ('Route / Supplier',  38),
    ('BoM Cost (€)',      13),
    ('Product Cost (€)',  15),
]

euro_fmt = '#,##0.00 "€"'
num_fmt  = '#,##0.00'


def _border():
    s = Side(style='thin', color='CFCFCF')
    return Border(left=s, right=s, top=s, bottom=s)


def write_excel(title, rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'BoM Overview'
    nc = len(COLS)
    lc = get_column_letter(nc)

    # Title row
    ws.merge_cells(f'A1:{lc}1')
    c = ws['A1']
    c.value = title or 'BoM Overview'
    c.font  = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    c.fill  = PatternFill('solid', fgColor=HEADER_DARK)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Header row
    hf = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    hb = PatternFill('solid', fgColor=HEADER_DARK)
    for ci, (cn, cw) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=cn)
        c.font = hf; c.fill = hb; c.border = _border()
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = cw
    ws.row_dimensions[2].height = 32

    # Data rows
    for ri, row in enumerate(rows, 3):
        rt = row['row_type']
        bg = (SUBCON_BG if rt == 'subcontracting' else
              OPS_BG    if rt in ('operations', 'operation_detail') else
              STRIPE    if ri % 2 == 0 else 'FFFFFF')
        fill = PatternFill('solid', fgColor=bg)
        bf   = Font(name='Arial', size=10)
        avail = row.get('availability', '') or ''
        afill = (PatternFill('solid', fgColor=AVAIL_RED)   if 'Not Available' in avail else
                 PatternFill('solid', fgColor=AVAIL_GREEN) if avail.startswith('Available') else
                 PatternFill('solid', fgColor=AVAIL_EST)   if 'Estimated' in avail else fill)

        vals = [row['code'], row['name'], row['quantity'], row['qty_unit'],
                row['ready_to_produce'], row['free_to_use'], row['on_hand'],
                avail, row['lead_time_days'], row['route'],
                row['bom_cost'], row['product_cost']]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = bf; c.border = _border()
            c.fill = afill if ci == 8 else fill
            c.alignment = Alignment(
                vertical='center', wrap_text=(ci in (2, 10)),
                horizontal='left' if ci in (1, 2, 10) else 'center')
            if ci in (3, 5, 6, 7) and isinstance(val, (int, float)):
                c.number_format = num_fmt
            if ci == 9 and isinstance(val, (int, float)):
                c.number_format = '0'
            if ci in (11, 12) and isinstance(val, (int, float)):
                c.number_format = euro_fmt
        ws.row_dimensions[ri].height = 18

    # Totals row
    ld = len(rows) + 2
    sr = ld + 2
    ws.merge_cells(f'A{sr}:J{sr}')
    c = ws[f'A{sr}']
    c.value = 'TOTAL'; c.font = Font(name='Arial', size=10, bold=True)
    c.fill = PatternFill('solid', fgColor='D9E1F2')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = _border()
    for gl in (get_column_letter(11), get_column_letter(12)):
        c = ws[f'{gl}{sr}']
        c.value = f'=SUM({gl}3:{gl}{ld})'
        c.font = Font(name='Arial', size=10, bold=True)
        c.fill = PatternFill('solid', fgColor='D9E1F2')
        c.number_format = euro_fmt
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = _border()
    ws.row_dimensions[sr].height = 20

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{lc}{ld}'
    wb.save(out_path)


# ── GUI ───────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title('Odoo BoM PDF → Excel')
        self.resizable(False, False)
        self.configure(bg='#1F3864')
        self.pdf_path = None
        self._build_ui()

    def _build_ui(self):
        tk.Label(self, text='Odoo BoM Overview', font=('Arial', 16, 'bold'),
                 bg='#1F3864', fg='white').pack(padx=20, pady=(18, 4))
        tk.Label(self, text='PDF  →  Excel Converter', font=('Arial', 11),
                 bg='#1F3864', fg='#AEC6CF').pack(pady=(0, 14))

        frame = tk.Frame(self, bg='white', bd=1, relief='solid')
        frame.pack(padx=20, pady=4, fill='x')
        self.path_var = tk.StringVar(value='No file selected…')
        tk.Label(frame, textvariable=self.path_var, font=('Arial', 9),
                 bg='white', fg='#444', anchor='w', width=50).pack(side='left', padx=10, pady=7)
        tk.Button(frame, text='Browse…', font=('Arial', 9),
                  command=self._browse, bg='#E8E8E8').pack(side='right', padx=6, pady=5)

        self.status = tk.Label(self, text='Select a BoM PDF to get started.',
                               font=('Arial', 10), bg='#1F3864', fg='#AEC6CF')
        self.status.pack(padx=20, pady=8)
        self.progress = ttk.Progressbar(self, mode='indeterminate', length=360)
        self.progress.pack(padx=20, pady=(0, 6))
        self.convert_btn = tk.Button(
            self, text='Convert to Excel', font=('Arial', 11, 'bold'),
            bg='#2E75B6', fg='white', activebackground='#1F5499',
            padx=24, pady=8, command=self._convert, state='disabled')
        self.convert_btn.pack(pady=(6, 20))

    def _browse(self):
        path = filedialog.askopenfilename(
            title='Select Odoo BoM Overview PDF',
            filetypes=[('PDF files', '*.pdf'), ('All files', '*.*')])
        if path:
            self.pdf_path = path
            self.path_var.set(Path(path).name)
            self.convert_btn.config(state='normal')
            self.status.config(text='Ready — click Convert to Excel.')

    def _convert(self):
        if not self.pdf_path:
            return
        self.convert_btn.config(state='disabled')
        self.progress.start(10)
        self.status.config(text='Parsing PDF…')
        self.update()
        try:
            title, rows = extract_rows(self.pdf_path)
            self.status.config(text=f'Found {len(rows)} rows — choose save location…')
            self.update()
            self.progress.stop()

            default_name = Path(self.pdf_path).stem + '.xlsx'
            out_path = filedialog.asksaveasfilename(
                title='Save Excel file as…', defaultextension='.xlsx',
                initialfile=default_name, filetypes=[('Excel files', '*.xlsx')])
            if out_path:
                self.progress.start(10)
                self.status.config(text='Writing Excel…')
                self.update()
                write_excel(title, rows, out_path)
                self.progress.stop()
                self.status.config(text=f'✔  Saved: {Path(out_path).name}')
                messagebox.showinfo('Conversion complete',
                    f'Excel file saved:\n\n{out_path}\n\n({len(rows)} rows exported)')
            else:
                self.progress.stop()
                self.status.config(text='Cancelled.')
        except Exception as e:
            self.progress.stop()
            messagebox.showerror('Conversion error', str(e))
            self.status.config(text='Error — see message box.')
        finally:
            self.convert_btn.config(state='normal')


if __name__ == '__main__':
    App().mainloop()
